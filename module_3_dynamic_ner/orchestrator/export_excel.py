import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def export_to_excel(data_records: list, output_filepath: str):
    """
    Hàm nhận mảng dữ liệu (List of Dictionaries) và vẽ file Excel chuẩn A4.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Muc_Luc_Ho_So"

    # --- 1. KHAI BÁO CÁC STYLE CSS CHO EXCEL ---
    font_header = Font(name='Times New Roman', bold=True, size=12)
    font_body = Font(name='Times New Roman', size=12)
    
    # Nền xám cho tiêu đề
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Kẻ viền đen mỏng cho mọi ô
    thin_border = Side(border_style="thin", color="000000")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Ép xuống dòng (Wrap Text)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # --- 2. VẼ DÒNG TIÊU ĐỀ ---
    headers = [
        "Số TT", 
        "Số, ký hiệu tài liệu", 
        "Ngày tháng văn bản", 
        "Tên loại hoặc trích yếu nội dung tài liệu", 
        "Tác giả tài liệu", 
        "Tờ số/Trang số"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border

    # --- 3. FIX CỨNG TỶ LỆ CỘT THEO KHUNG A4 TRỌC PHÚ ---
    # Tổng độ rộng khoảng 110-120 đơn vị là vừa khít trang A4 khổ dọc
    col_widths = {
        'A': 6,   # STT
        'B': 18,  # Số, ký hiệu
        'C': 15,  # Ngày tháng
        'D': 45,  # Tên loại (Dài nhất)
        'E': 25,  # Tác giả
        'F': 10   # Tờ số
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- 4. ĐỔ DỮ LIỆU & FORMAT TỪNG DÒNG ---
    for idx, record in enumerate(data_records, 1):
        row_data = [
            idx,
            record.get('so_ky_hieu_tai_lieu', ''),
            record.get('ngay_thang_van_ban', ''),
            record.get('ten_loai_tai_lieu', ''),
            record.get('tac_gia', ''),
            record.get('to_so_trang_so', '')
        ]
        ws.append(row_data)
        
        # Lấy dòng vừa mới thêm vào để format
        current_row = ws[ws.max_row]
        for cell in current_row:
            cell.font = font_body
            cell.border = border
            
        # Căn trái riêng cho cột Tên loại (cột D), còn lại căn giữa
        current_row[0].alignment = align_center
        current_row[1].alignment = align_center
        current_row[2].alignment = align_center
        current_row[3].alignment = align_left
        current_row[4].alignment = align_center
        current_row[5].alignment = align_center

    # --- 5. LƯU FILE ---
    # Đảm bảo thư mục tồn tại
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    wb.save(output_filepath)
    print(f"\n✅ Đã xuất báo cáo Excel thành công tại: {output_filepath}")